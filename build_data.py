import json
from pathlib import Path
import openpyxl

BOOK = Path("data/원본엑셀.xlsx")
OUT = Path("data/result.json")

def norm(v):
    return str(v).strip().lower().replace(" ","").replace("_","")

def num(v):
    if v is None: return None
    try: return int(float(v))
    except: return None

try:
    wb = openpyxl.load_workbook(BOOK, data_only=True, read_only=True)
    if "FINAL_자동발급" not in wb.sheetnames:
        raise ValueError("FINAL_자동발급 시트가 없습니다.")
    if "로또데이터" not in wb.sheetnames:
        raise ValueError("로또데이터 시트가 없습니다.")

    ws = wb["FINAL_자동발급"]
    rows = list(ws.iter_rows(values_only=True))
    hi = None
    for i,row in enumerate(rows[:30]):
        h=[norm(x) for x in row]
        if "순번" in h and all(f"n{j}" in h for j in range(1,7)):
            hi=i; header=h; break
    if hi is None:
        raise ValueError("FINAL_자동발급의 순번/N1~N6 헤더를 찾지 못했습니다.")

    rc=header.index("순번")
    nc=[header.index(f"n{j}") for j in range(1,7)]
    numbers=None
    for row in rows[hi+1:]:
        if num(row[rc])==1:
            numbers=[num(row[c]) for c in nc]
            break
    if numbers is None:
        raise ValueError("순번=1 행을 찾지 못했습니다.")
    if len(numbers)!=6 or any(n is None or n<1 or n>45 for n in numbers):
        raise ValueError("1순위 번호가 올바르지 않습니다.")
    if len(set(numbers))!=6:
        raise ValueError("1순위 번호에 중복이 있습니다.")

    wd=wb["로또데이터"]
    drows=list(wd.iter_rows(values_only=True))
    hidx=None; col=None
    for i,row in enumerate(drows[:30]):
        vals=[str(x).strip() if x is not None else "" for x in row]
        if "회차" in vals:
            hidx=i; col=vals.index("회차"); break
    if col is None:
        raise ValueError("로또데이터의 회차 열을 찾지 못했습니다.")
    latest=max(num(row[col]) for row in drows[hidx+1:] if num(row[col]) is not None)

    result={"ok":True,"latestRound":latest,"nextRound":latest+1,"numbers":numbers,
            "sourceSheet":"FINAL_자동발급","rank":1,"validated":True}
except Exception as e:
    result={"ok":False,"error":str(e)}

OUT.write_text(json.dumps(result,ensure_ascii=False,indent=2),encoding="utf-8")
if not result["ok"]:
    raise SystemExit(result["error"])
print(result)
